from django.shortcuts import redirect, get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from pomi.models.stundent import Student
from pomi.models.whatsappUser import WhatsAppUserStudent
from django.contrib import messages
from pomi.views.navbar import get_navbar_context
from pomi.models.ticket import Ticket
import csv
import json
import openpyxl
from io import BytesIO

@login_required
def manage_student(request):
    """
    CRUD compacto:
    - GET   -> lista alumnos
    - POST  -> create / update / delete (campo hidden 'action')
    """
    if request.method == 'POST':
        action = request.POST.get('action', 'create')
        code_upc = request.POST.get('code_upc', '').strip().lower()

        if action == 'delete':
            obj = get_object_or_404(Student, code_upc=code_upc)
            obj.delete()
            messages.success(request, f"Alumno {code_upc} eliminado.")
            return redirect('pomi:alumnos')

        # datos comunes create/update
        first_names = request.POST.get('first_names', '').strip()
        full_names  = request.POST.get('full_names', '').strip()
        career      = request.POST.get('career', '').strip()
        is_active   = request.POST.get('is_active') in ['on', 'true', 'True', '1']

        if not code_upc or not first_names or not full_names:
            messages.error(request, "Código UPC, Nombres y Apellidos son obligatorios.")
            return redirect('pomi:alumnos')

        if action == 'update':
            obj = get_object_or_404(Student, code_upc=code_upc)
            obj.first_names = first_names
            obj.full_names  = full_names
            obj.career      = career
            obj.is_active   = is_active
            obj.save()
            messages.success(request, f"Alumno {code_upc} actualizado.")
        else:
            Student.objects.create(
                code_upc=code_upc,
                first_names=first_names,
                full_names=full_names,
                career=career,
                is_active=is_active,
            )
            messages.success(request, f"Alumno {code_upc} creado.")

        return redirect('pomi:alumnos')

    # GET
    # Obtener parámetro de búsqueda
    query = request.GET.get('q', '').strip()
    
    # Filtrar alumnos
    if hasattr(Student, 'created_at'):
        alumnos_list = Student.objects.all().order_by('-created_at')
    else:
        alumnos_list = Student.objects.all().order_by('code_upc')
    
    # Aplicar búsqueda si existe
    if query:
        from django.db.models import Q
        alumnos_list = alumnos_list.filter(
            Q(code_upc__icontains=query) |
            Q(first_names__icontains=query) |
            Q(full_names__icontains=query) |
            Q(career__icontains=query)
        )
    
    # Paginación
    paginator = Paginator(alumnos_list, 10)  # 10 alumnos por página
    page_number = request.GET.get('page', 1)
    
    try:
        alumnos = paginator.page(page_number)
    except PageNotAnInteger:
        alumnos = paginator.page(1)
    except EmptyPage:
        alumnos = paginator.paginator.page(paginator.num_pages)
    
    ctx = get_navbar_context(request, current_page='students')
    ctx.update({
        'page_title': 'Alumnos',
        'page_subtitle': 'Gestión de estudiantes',
        'alumnos': alumnos,
        'total_alumnos': paginator.count,
    })
    return render(request, 'alumnos/manage.html', ctx)

@login_required
def export_students_csv(request):
    """
    Exporta la lista de alumnos a un archivo CSV
    """
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="alumnos.csv"'
    response.write('\ufeff')  # BOM para UTF-8
    
    writer = csv.writer(response)
    writer.writerow(['Código UPC', 'Nombres', 'Apellidos', 'Carrera', 'Estado'])
    
    alumnos = Student.objects.all().order_by('code_upc')
    for alumno in alumnos:
        writer.writerow([
            alumno.code_upc,
            alumno.first_names,
            alumno.full_names,
            alumno.career or '',
            'Activo' if alumno.is_active else 'Inactivo'
        ])
    
    return response

@login_required
def download_template(request):
    """
    Descarga una plantilla CSV de ejemplo para carga masiva
    """
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="plantilla_alumnos.csv"'
    response.write('\ufeff')  # BOM para UTF-8
    
    writer = csv.writer(response)
    writer.writerow(['code_upc', 'first_names', 'full_names', 'career', 'is_active'])
    writer.writerow(['U202012345', 'Juan Carlos', 'Pérez García', 'Ing. de Sistemas', 'true'])
    writer.writerow(['U202012346', 'María Elena', 'López Martínez', 'Ing. de Software', 'true'])
    writer.writerow(['U202012347', 'Pedro Luis', 'Rodríguez Silva', 'C.C.', 'false'])
    
    return response

@login_required
@require_POST
def process_bulk_file(request):
    """
    Procesa el archivo CSV/XLSX y devuelve los datos validados
    """
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'No se recibió ningún archivo'})
        
        file = request.FILES['file']
        mode = request.POST.get('mode', 'add')
        
        # Determinar tipo de archivo y procesar
        if file.name.endswith('.csv'):
            data = process_csv_file(file)
        elif file.name.endswith('.xlsx'):
            data = process_xlsx_file(file)
        else:
            return JsonResponse({'success': False, 'error': 'Formato de archivo no soportado'})
        
        # Validar datos
        validated_data = validate_student_data(data, mode)
        
        return JsonResponse({
            'success': True,
            'data': validated_data
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_POST
def confirm_bulk_load(request):
    """
    Confirma y ejecuta la carga masiva de alumnos
    """
    try:
        body = json.loads(request.body)
        mode = body.get('mode', 'add')
        data = body.get('data', [])
        
        if not data:
            return JsonResponse({'success': False, 'error': 'No hay datos para procesar'})
        
        if mode == 'replace':
            # Eliminar todos los alumnos existentes
            Student.objects.all().delete()
            action = 'reemplazados'
        else:
            action = 'agregados'
        
        # Insertar nuevos alumnos
        created_count = 0
        updated_count = 0
        
        for row in data:
            code_upc = row.get('code_upc', '').strip()
            
            # Verificar si existe
            student, created = Student.objects.update_or_create(
                code_upc=code_upc,
                defaults={
                    'first_names': row.get('first_names', '').strip(),
                    'full_names': row.get('full_names', '').strip(),
                    'career': row.get('career', '').strip(),
                    'is_active': row.get('is_active', True)
                }
            )
            
            if created:
                created_count += 1
            else:
                updated_count += 1
        
        message = f'{created_count} alumno(s) creado(s)'
        if updated_count > 0:
            message += f', {updated_count} actualizado(s)'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'created': created_count,
            'updated': updated_count
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def process_csv_file(file):
    """
    Procesa un archivo CSV y retorna lista de diccionarios
    """
    data = []
    decoded_file = file.read().decode('utf-8-sig').splitlines()
    reader = csv.DictReader(decoded_file)
    
    for row in reader:
        data.append(row)
    
    return data

def process_xlsx_file(file):
    """
    Procesa un archivo XLSX y retorna lista de diccionarios
    """
    data = []
    wb = openpyxl.load_workbook(BytesIO(file.read()))
    ws = wb.active
    
    # Obtener encabezados de la primera fila
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    # Leer datos
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for idx, value in enumerate(row):
            if idx < len(headers):
                row_data[headers[idx]] = value
        data.append(row_data)
    
    return data

def validate_student_data(data, mode):
    """
    Valida los datos de estudiantes
    """
    validated = []
    existing_codes = set(Student.objects.values_list('code_upc', flat=True))
    
    for idx, row in enumerate(data):
        errors = []
        
        # Obtener valores
        code_upc = str(row.get('code_upc', '')).strip() if row.get('code_upc') else ''
        first_names = str(row.get('first_names', '')).strip() if row.get('first_names') else ''
        full_names = str(row.get('full_names', '')).strip() if row.get('full_names') else ''
        career = str(row.get('career', '')).strip() if row.get('career') else ''
        is_active_raw = str(row.get('is_active', 'true')).strip().lower() if row.get('is_active') else 'true'
        
        # Validar código UPC (obligatorio)
        if not code_upc:
            errors.append('Código UPC requerido')
        elif len(code_upc) > 10:
            errors.append('Código UPC muy largo (máx 10 caracteres)')
        elif mode == 'add' and code_upc in existing_codes:
            errors.append('Código UPC ya existe en la BD')
        
        # Validar nombres (obligatorio)
        if not first_names:
            errors.append('Nombres requeridos')
        elif len(first_names) > 100:
            errors.append('Nombres muy largos (máx 100 caracteres)')
        
        # Validar apellidos (obligatorio)
        if not full_names:
            errors.append('Apellidos requeridos')
        elif len(full_names) > 100:
            errors.append('Apellidos muy largos (máx 100 caracteres)')
        
        # Validar carrera (opcional)
        if career and len(career) > 50:
            errors.append('Carrera muy larga (máx 50 caracteres)')
        
        # Validar y parsear is_active
        is_active = parse_boolean(is_active_raw)
        
        validated.append({
            'code_upc': code_upc,
            'first_names': first_names,
            'full_names': full_names,
            'career': career,
            'is_active': is_active,
            'errors': errors
        })
    
    return validated

def parse_boolean(value):
    """
    Convierte varios formatos a boolean
    """
    if isinstance(value, bool):
        return value
    
    value_lower = str(value).lower().strip()
    
    true_values = ['true', '1', 'si', 'sí', 'yes', 'activo', 'active', 't', 'y']
    false_values = ['false', '0', 'no', 'inactivo', 'inactive', 'f', 'n']
    
    if value_lower in true_values:
        return True
    elif value_lower in false_values:
        return False
    else:
        return True  # Por defecto True

# APIs
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from pomi.apis.studentSerializer import VerifyStudentSerializer
from pomi.apis.studentServices import update_or_create_WhatsAppUser

class VerifyStudentAPI(APIView):
    
    """
    API para verificar si un estudiante existe por su código
    y asociar/actualizar su número de WhatsApp
    """
    
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        print(request.data)
        serializer = VerifyStudentSerializer(data=request.data)
        if not serializer.is_valid():
            print(serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        # Obtener datos validados
        student = serializer.validated_data['student']
        phone = serializer.validated_data['phone']
        
        # Crear o actualizar WhatsAppUser
        try:
            result = update_or_create_WhatsAppUser(student, phone)
        except Exception as e:
            print({'error': str(e)})
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # Preparar respuesta
        response_data = {
            'success': True,
            'message': f'Estudiante verificado y WhatsApp {result["state"]}',
            'data': {
                'code': student.code_upc,
                'names': student.first_names,
                'last_names': student.full_names,
                'whatsapp_status': result["state"],
                'activo': student.is_active
            }
        }
        
        return Response(response_data, status=status.HTTP_200_OK)
       
